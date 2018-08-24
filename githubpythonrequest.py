#pip install xlrd
#安装excel的interpreter
#pip install openpyxl
#安装使用excel的包
#pip install urllib
#安装request的语句，目前python3已经包含
# -*- coding: utf-8 -*-
import random
import urllib.request
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter




head_connection = ['Keep-Alive','close']
# 开启 keep-alive模式，请求处理完毕会被断掉
head_accept = ['text/html,application/xhtml+xml,*/*']
head_accept_language = ['zh-CN,fr-FR;q=0.5','en-US,en;q=0.8,zh-Hans-CN;q=0.5,zh-Hans;q=0.3']
head_user_agent = ['Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko',
                    'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.95 Safari/537.36',
                    'Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; rv:11.0) like Gecko)',
                    'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1',
                    'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070309 Firefox/2.0.0.3',
                    'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12',
                    'Opera/9.27 (Windows NT 5.2; U; zh-cn)',
                    'Mozilla/5.0 (Macintosh; PPC Mac OS X; U; en) Opera 8.0',
                    'Opera/8.0 (Macintosh; PPC Mac OS X; U; en)',
                    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.12) Gecko/20080219 Firefox/2.0.0.12 Navigator/9.0.0.6',
                    'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)',
                    'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)',
                    'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.2; .NET4.0C; .NET4.0E)',
                    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Maxthon/4.0.6.2000 Chrome/26.0.1410.43 Safari/537.1 ',
                    'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; InfoPath.2; .NET4.0C; .NET4.0E; QQBrowser/7.3.9825.400)',
                    'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:21.0) Gecko/20100101 Firefox/21.0 ',
                    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/21.0.1180.92 Safari/537.1 LBBROWSER',
                    'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; BIDUBrowser 2.x)',
                    'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.11 TaoBrowser/3.0 Safari/536.11']


def get_header() :
    header={
        'Connect':head_connection[random.randrange(0,len(head_connection))],
        'Accept': head_accept[0],
        'Accept-Language' : head_accept_language[random.randrange(0,len(head_accept_language))],
        'User-Agent': head_user_agent[random.randrange(0,len(head_user_agent))],
    }
    return header

#上面是构建请求头， 伪造多个浏览器


#下面为使用代理和获取网页内容

def get_html(url):
    header = get_header()
    proxyIP=['183.147.209.69:9080']
    #设置代理ip和端口
    proxyhandler= urllib.request.ProxyHandler({'http':proxyIP[0]})
    #设置代理
    opener=urllib.request.build_opener(proxyhandler)
    #创建opener对象
    head=[]
    for key,value in header.items():
        aa=(key,value)
        head.append(aa)
    opener.addheaders=head
    #把隐匿ip放入head里
    send_head=opener.open(url)
    #向服务器发送请求头，并获取反应结果
    data=send_head.read().decode('utf-8')
    return data

# import urllib.request
# url = r'https://movie.douban.com/top250?start=25&filter='
# res = urllib.request.urlopen(url)
# html = res.read().decode('utf-8')
# print(html)
#获取网页的html


datalist = []
#创建存放数据列表
def qingxi(data):
    #定义清理函数，利用正则表达式提取数据
    reg=re.compile('<div class="item">.*?<img width="100" alt="(.*?)".*?<p class="">.*?导演: (.*?) .*?<span class="rating_num" property="v:average">(.*?)</span>.*?<span>(.*?)人评价</span>.*?</div>',re.S)
    lists=re.findall(reg,data)
    datalist.extend(lists)


i=0
while i <= 225:
    print('获取第',(i/25+1),'页')
    url='https://movie.douban.com/top250?start='+str(i)+'&filter='
    data=get_html(url)
    qingxi(data)
    i+=25
#找到网页的规律

print(datalist)

def saveExcel():
    wb=Workbook()
    sheet1=wb.create_sheet('豆瓣电影信息',0)
    sheet1.cell(1,1).value='电影名称'
    sheet1.cell(1,2).value='导演'
    sheet1.cell(1,3).value='评分'
    sheet1.cell(1,4).value='评论人数'
    #创建excel，和第一页，和第一行信息，注意cell是从1开始
    for i in range(len(datalist)):
        for j in range(len(datalist[i])):
            sheet1.cell(i+2,j+1).value=datalist[i][j]
    wb.save('doubanresult.xlsx')
    #使用openpyxl，可以直接保存为xlsx
saveExcel()